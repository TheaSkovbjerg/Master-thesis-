from nicegui import classes, events, ui
from pydantic import BaseModel 
from openai import AsyncClient, OpenAI
from pypdf import PdfReader
from io import BytesIO
from typing import Optional
from typing import Literal
import pdfplumber
from openpyxl import Workbook, load_workbook
import os

all_candidates = []

client = AsyncClient(api_key='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')

input_area = ui.textarea(label='PDF text').classes('w-full')
output_area = ui.textarea(label='Extracted text').classes('w-full')

class CandidateInfo(BaseModel):
    name: Optional[str] = None
    bachelor_gpa: Optional[list[str]] = None
    bachelor_gpa_snippets: Optional[list[str]] = None
    bachelor_education: Optional[list[str]] = None
    bachelor_education_snippets: Optional[list[str]] = None
    master_gpa: Optional[list[str]] = None
    master_gpa_snippets: Optional[list[str]] = None
    master_education: Optional[list[str]] = None
    master_education_snippets: Optional[list[str]] = None
    exchange_gpa: Optional[list[str]] = None
    exchange_gpa_snippets: Optional[list[str]] = None
    exchange_education: Optional[list[str]] = None
    exchange_education_snippets: Optional[list[str]] = None
    student_job: Optional[list[str]] = None
    student_job_snippets: Optional[list[str]] = None
    persona: Optional[str] = None
    persona_reason: Optional[str] = None
    persona_snippets: Optional[list[str]] = None

start_prompt = """
Choose exactly one of these five personas based on evidence in the CV, cover letter and recommendation (if one of these have been uploaded).
"""

persona_1_prompt = """
This is Sophia Rødkjær. In her cover letter/CV, she, among other things, writes in her cover letter: "I am passionate about leveraging technology to drive business transformation and create innovative solutions that deliver tangible value. With a strong fundament in digital transformation, I have successfully helped C-level planning technology-enabled business models in my student job."
Sophia has a bachelors degree in Business Administration and Digital Management from CBS.
Sophia has a masters degree in Business Administration and Digital Business from CBS.

Sophie is classified as Persona 1 - Tech Strategist because she mentions the keywords digital transformation, technology-enabled business models. She is also classified as Persona 1 - Tech Strategist based on her education in Business Administration & Digital Business from CBS, which is one of the key programs targeted for the Tech Strategist persona.

Persona 1 - Tech Strategist is defined by the following criteria:
[The candidate has one or more of the following keywords: Strategic problem solving, Digital transformation, Technology-enabled business models, C-level advisory, Operating model design, Business case development, Technology strategy.
AND the candidate has attended one of these primary Target Universities: CBS, Aarhus University (AU), University of Copenhagen (KU).
OR the candidate has attended one of these secondary Universities: DTU (tech management tracks), Lund University.
AND these are the Key Programs to Target: Business Administration & Digital Business (CBS), Strategy, Organization & Leadership (CBS), Management of Innovation & Business Development (CBS), Economics (KU / AU), International Business (CBS / AU), Digital Business Management (AU).]
"""

persona_2_prompt = """
This is Mikkel Jensen. In his cover letter/CV he mentions working with project management in his student job. Further, he coordinated communication to and with the stakeholder.
Mikkel holds a bachelors degree in Economics and Business Administration from Aarhus University.
Mikkel holds a masters degree in Digital Business Management from Aarhus University.

Mikkel is classified as persona 2 - Transformation Orchestrator (PMO) because of his experience with project management and stakeholder coordination in his cover letter as well as his education in Digital Business Management from Aarhus University, which is one of the key programs targeted for the Transformation Orchestrator persona.

Persona 2 - Transformation Orchestrator (PMO) is defined by the following criteria:
[The candidate has one or more of the following keywords: Program management, Execution & value realization, Governance & operating models, Change enablement, Stakeholder coordination, Implementation roadmap design.
AND the candidate has attended one of these primary Target Universities: CBS, Aarhus University (AU), DTU.
OR the candidate has attended one of these secondary Universities: ITU, AAU.
AND these are the Key Programs to Target: Digital Business Management (AU), Business Administration & Information Systems (CBS), Management of Innovation & Business Development (CBS), Operations & Supply Chain programs, Engineering with business specialization (DTU).]
"""

persona_3_prompt = """
This is Laura Hansen. In her cover letter/CV, she writes: "I have a strong passion for data and analytics, and I am eager to apply my skills in analytics interpretation and GenAI applications to drive informed decision-making and business insights. Previously, at my student job, I have successfully completed projects that involved analyzing complex datasets, developing predictive models, and communicating insights to both technical and non-technical stakeholders."
Laura has a bachelors degree in Economics and Business Administration from CBS and a masters degree in Business Intelligence from AU.  

Laura is classified as Persona 3 - Tech Translator because she mentions the keywords analytics interpretation and GenAI applications in her cover letter. Laura also has a masters degree in Business Intelligence from Aarhus University, which is one of the key programs targeted for the Tech Translator persona.

Persona 3 - Tech Translator is defined by the following criteria:
[Key Words: Data literacy, AI enablement, Business-IT alignment, Systems & architecture understanding, Analytics interpretation, Use-case development, GenAI applications. 
Primary Target Universities: ITU, DTU, AAU.
Secondary / Feeder Universities: CBS (Data / Digital tracks), KU (Quantitative / Data tracks).
Key Programs to Target: Business Intelligence (CBS / AU), Business Administration & Data Science (CBS), Computer Science (ITU / DTU / AAU), Data Science programs, Analytics & Information Systems programs, IT, Communication & Organisation.]
"""

persona_4_prompt = """
This is Patricia. In her cover letter/CV she mentions working with technical solution development in her student job. She also mentions working with machine learning and data analysis in her thesis. 
Patricia holds a bachelors degree in Software Engineering from SDU.
Patricia has a masters degree in Engeneering from SDU.

Patricia is classified as Persona 4 - Applied Technical Specialist because of her experience with technical solution development, machine learning and data analysis in her cover letter and CV. Her education in Software Engineering from SDU also supports this classification, as SDU is listed as a secondary university for the Applied Technical Specialist persona.

Persona 4 - Applied Technical Specialist is defined by the following criteria:
[Key Words: Advanced quantitative methods, Machine learning & modelling, Algorithmic thinking, Statistical analysis, Optimization & simulation, Technical solution development, Systems engineering.
Primary Target Universities: DTU, ITU, AAU, KU (Science Faculty), Lund University (Engineering).
Secondary / Feeder Universities: SDU, International technical programs.
Key Programs to Target: Computer Science, Applied Mathematics, Physics, Statistics, Engineering (software, systems, industrial, energy), Machine Learning & AI programs.]
"""

persona_5_prompt = """
This is Emilie Jensen. In her cover letter, she writes: "I have a strong passion for public sector strategy. In my student job, I have successfully completed projects that involved using my regulatory knowledge to help provide the public sector with industry-specific solutions." 
Emilie has a bachelors degree in Political Science from KU.
Emilie has a masters degree in Political Science from KU.

Emilie is classified as Persona 5 - Industry / Functional Specialist because she mentions the keywords sector strategy and regulatory knowledge in her cover letter. Her education in Political Science from KU also supports this classification, as KU is listed as a primary target university for the Industry / Functional Specialist persona.

Persona 5 - Industry / Functional Specialist is defined by the following criteria:
[Key Words: Industry expertise, Regulatory knowledge, Value chain insight, Commercial awareness, Sector strategy, Domain-specific analytics, Strategy execution.
Primary Target Universities: KU (Life Science, Public Sector), DTU (Energy, Engineering), CBS (Financial Services, Commercial), Aarhus University (AU), Lund University.
Secondary / Feeder Universities: RUC, SDU, AAU.
Key Programs to Target: Applied Economics & Finance (KU), Finance & Strategic Management (CBS), Life Science / Biomed programs (e.g. Biotechnology, Bioinformatics, Life Science Engineering, Pharmacy etc. at KU / DTU / Lund), Energy & Engineering programs (DTU / AAU), Supply Chain & Operations programs, Public Policy / Political Science (KU / AU).]
"""

end_prompt = """
Based on the examples of the five personas and the criteria for each persona, analyze the CV, cover letter and recommendation (if uploaded) and determine which single persona best fits the candidate. Let's think step by step.
Return only the single best fitting persona.
Also return a short reason and exact snippets from the CV supporting the choice.
If evidence is insufficient, return null.
"""

async def send_input():
    try:
        t = input_area.value or ''

        print("TEXT LENGTH:", len(t)) 
        print("FULL TEXT:", repr(t))

        completion = await client.beta.chat.completions.parse(
            model="gpt-4.1",
            messages=[
                {"role": "system", "content": (
                    "Extract all bachelor entries and bachelor GPA values from the text. "
                    "Extract all master entries and master GPA values from the text. "
                    "Extract all exchange entries and exchange GPA values from the text (if any). "
                    "Exchange may be written as 'exchange', 'exchange semester', 'exchange program', 'study abroad', or similar. "
                    "Student job refers to the most recent relevant student job experience mentioned in the text, and may be written as 'student job', 'student assistant', 'internship', or similar. "
                    "GPA may be written as 'grade average', 'average grade', or similar. "
                    "The GPA may appear on the same line or the next line after the label. "
                    "Return the GPA's exactly as written. "
                    "For each extracted value, also return the exact snippet from the text that supports it. "
                    "Return all entries you find, not just the most recent one. "
                    "education_snippets should be short exact substrings from the text."
                    "If a value is not explicitly present, return null. Do not guess. "
                    + start_prompt
                    + persona_1_prompt
                    + persona_2_prompt
                    + persona_3_prompt
                    + persona_4_prompt
                    + persona_5_prompt
                    + end_prompt
                )},
                {"role": "user", "content": t},
            ],
            response_format=CandidateInfo,
        )

        npa = completion.choices[0].message.parsed

        global all_candidates
        all_candidates.append(npa)

        print(npa)

        bachelor_gpa_snips = npa.bachelor_gpa_snippets or []
        bachelor_education_snips = npa.bachelor_education_snippets or []
        master_gpa_snips = npa.master_gpa_snippets or []
        master_education_snips = npa.master_education_snippets or []
        exchange_gpa_snips = npa.exchange_gpa_snippets or []
        exchange_education_snips = npa.exchange_education_snippets or []
        student_job_snips = npa.student_job_snippets or []
        persona_snips = npa.persona_snippets or []

        snippets = bachelor_gpa_snips + bachelor_education_snips + master_gpa_snips + master_education_snips + exchange_gpa_snips + exchange_education_snips + student_job_snips + persona_snips

        output_area.value = (
            f"name={npa.name}\n"
            f"master_education={npa.master_education}\n"
            f"master_gpa={npa.master_gpa}\n"
            f"bachelor_education={npa.bachelor_education}\n"
            f"bachelor_gpa={npa.bachelor_gpa}\n"
            f"exchange_education={npa.exchange_education}\n"
            f"student_job={npa.student_job}\n"
            f"persona={npa.persona}\n"
            f"persona_reason={npa.persona_reason}"
            )
        
        highlighted_output.set_content(
            highlight_all(t, snippets)
            )

    except Exception as e:
        print("ERROR:", e)
        output_area.value = str(e)

ui.button("Extract", on_click=send_input)


#Upload pdf file to Nice Gui
async def uploads(e: events.UploadEventArguments):
    pdf_bytes = await e.file.read()
    text = ''
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text += page.extract_text() or ''
    
    text = text or 'No text found in PDF.'
    
    if input_area.value:
        input_area.value += '\n\n----- Next document -----\n\n' + text
    else:
        input_area.value = text

ui.upload(on_upload=uploads, auto_upload=True, multiple=True).props('accept=.pdf')

markdown = ui.markdown('Choose a PDF file!')

#Add second area for highlighting nice gui
# create UI element
highlighted_output = ui.html().classes('w-full')

def highlight_all(t, snippets):
    for s in snippets:
        if s:
            t = t.replace(s, f'<mark>{s}</mark>')
    return t.replace('\n', '<br>')

def save_to_excel():
    if not all_candidates: ui.notify('No data'); return

    path = '/Users/danicahennelly/Speciale UDENFOR iCloud/TEST.xlsx'
    wb = load_workbook(path) if os.path.exists(path) else Workbook()
    
    #Del det op i forskellige faner
    for c in all_candidates:
        persona = c.persona or 'Unknown'
        # clean sheet name (Excel doesn't like long names)
        sheet_name = persona[:31]
        # create or get sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            # add headers ONLY when sheet is created
            ws.append([
                'name', 'master_education', 'master_gpa',
                'bachelor_education', 'bachelor_gpa',
                'exchange_education', 'exchange_gpa',
                'student_job', 'persona', 'persona_reason'
            ])

            # append row
            ws.append([
                c.name[0] if isinstance(c.name, list) else c.name or '',
                ', '.join(c.master_education or []),
                ', '.join(c.master_gpa or []),
                ', '.join(c.bachelor_education or []),
                ', '.join(c.bachelor_gpa or []),
                ', '.join(c.exchange_education or []),
                ', '.join(c.exchange_gpa or []),
                ', '.join(c.student_job or []),
                c.persona or '',
                (c.persona_reason or '').replace('. ', '.\n')
            ])
            
    #Dette opretter faner
    if ws.max_row == 1:
        ws.append(['name','master_education','master_gpa','bachelor_education','bachelor_gpa','exchange_education','exchange_gpa','student_job','persona','persona_reason'])

    for c in all_candidates:
        ws.append([
            c.name[0] if isinstance(c.name, list) else c.name or '',
            ', '.join(c.master_education or []),
            ', '.join(c.master_gpa or []),
            ', '.join(c.bachelor_education or []),
            ', '.join(c.bachelor_gpa or []),
            ', '.join(c.exchange_education or []),
            ', '.join(c.exchange_gpa or []),
            ', '.join(c.student_job or []),
            c.persona or '',
            (c.persona_reason or '').replace('. ', '.\n')
        ])

    wb.save(path); ui.notify(f'Saved {len(all_candidates)}'); all_candidates.clear()

ui.button("Save to Excel", on_click=save_to_excel)


ui.run()